"""Extract sheet-level structural data from openpyxl worksheets."""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from exceldiff.parsing.cell_parser import parse_cell
from exceldiff.parsing.models import (
    CellData,
    ConditionalFormatData,
    DataValidationData,
    SheetData,
)


def parse_sheet(ws: Worksheet) -> SheetData:
    """Parse an openpyxl worksheet into a SheetData snapshot."""
    # Merged cells
    merged = tuple(sorted(str(r) for r in ws.merged_cells.ranges))

    # Column widths (non-default only)
    column_widths: dict[str, float] = {}
    for col_letter, dim in ws.column_dimensions.items():
        if dim.width is not None and dim.width != 8.0:
            column_widths[col_letter] = dim.width

    # Row heights (non-default only)
    row_heights: dict[int, float] = {}
    for row_num, dim in ws.row_dimensions.items():
        if dim.height is not None and dim.height != 15.0:
            row_heights[row_num] = dim.height

    # Data validations
    data_validations: list[DataValidationData] = []
    if ws.data_validations and ws.data_validations.dataValidation:
        for dv in ws.data_validations.dataValidation:
            cells = tuple(sorted(str(r) for r in dv.cells.ranges)) if dv.cells else ()
            data_validations.append(
                DataValidationData(
                    type=dv.type,
                    formula1=str(dv.formula1) if dv.formula1 else None,
                    formula2=str(dv.formula2) if dv.formula2 else None,
                    cells=cells,
                )
            )

    # Conditional formatting
    conditional_formats: list[ConditionalFormatData] = []
    for cf_range, rules in ws.conditional_formatting._cf_rules:
        for rule in rules:
            formula_text = None
            if rule.formula and len(rule.formula) > 0:
                formula_text = str(rule.formula[0])
            conditional_formats.append(
                ConditionalFormatData(
                    rule_type=rule.type or "",
                    formula=formula_text,
                    cells=str(cf_range),
                    priority=rule.priority or 0,
                )
            )

    # Parse all cells
    cells: dict[str, CellData] = {}
    for row in ws.iter_rows():
        for cell in row:
            parsed = parse_cell(cell)
            if parsed is not None:
                cells[parsed.coordinate] = parsed

    # Auto filter
    auto_filter = None
    if ws.auto_filter and ws.auto_filter.ref:
        auto_filter = ws.auto_filter.ref

    # Print area
    print_area = None
    if ws.print_area:
        print_area = str(ws.print_area)

    # Sheet protection
    sheet_protection = bool(ws.protection.sheet) if ws.protection else False

    return SheetData(
        name=ws.title,
        dimensions=ws.dimensions or "",
        max_row=ws.max_row or 0,
        max_column=ws.max_column or 0,
        merged_cells=merged,
        freeze_panes=str(ws.freeze_panes) if ws.freeze_panes else None,
        column_widths=column_widths,
        row_heights=row_heights,
        data_validations=tuple(data_validations),
        conditional_formats=tuple(conditional_formats),
        cells=cells,
        auto_filter=auto_filter,
        print_area=print_area,
        sheet_protection=sheet_protection,
    )
