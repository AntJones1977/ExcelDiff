"""Tests for annotated Excel output renderer."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from exceldiff.comparison.engine import compare
from exceldiff.renderers.excel import render


class TestExcelRenderer:
    def test_generates_xlsx_file(self, base_path, modified_formulas_path, tmp_path):
        diff = compare(str(base_path), str(modified_formulas_path))
        out = str(tmp_path / "annotated.xlsx")
        result_path = render(diff, out)
        assert Path(result_path).exists()

    def test_output_is_valid_xlsx(self, base_path, modified_formulas_path, tmp_path):
        diff = compare(str(base_path), str(modified_formulas_path))
        out = str(tmp_path / "annotated.xlsx")
        render(diff, out)
        wb = openpyxl.load_workbook(out)
        assert len(wb.sheetnames) > 0
        wb.close()

    def test_contains_summary_sheet(self, base_path, modified_formulas_path, tmp_path):
        diff = compare(str(base_path), str(modified_formulas_path))
        out = str(tmp_path / "annotated.xlsx")
        render(diff, out)
        wb = openpyxl.load_workbook(out)
        assert "_ExcelDiff_Summary" in wb.sheetnames
        assert wb.sheetnames[0] == "_ExcelDiff_Summary"
        wb.close()

    def test_summary_has_stats(self, base_path, modified_formulas_path, tmp_path):
        diff = compare(str(base_path), str(modified_formulas_path))
        out = str(tmp_path / "annotated.xlsx")
        render(diff, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["_ExcelDiff_Summary"]
        # Check that summary content exists
        cell_values = [ws.cell(r, 1).value for r in range(1, 20) if ws.cell(r, 1).value]
        assert "ExcelDiff Report" in cell_values
        assert "Formula diffs" in cell_values
        wb.close()

    def test_changed_cells_have_comments(self, base_path, modified_formulas_path, tmp_path):
        diff = compare(str(base_path), str(modified_formulas_path))
        out = str(tmp_path / "annotated.xlsx")
        render(diff, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Summary"]
        # B2 formula changed — should have a comment
        assert ws["B2"].comment is not None
        assert "ExcelDiff" in ws["B2"].comment.text
        wb.close()

    def test_added_sheet_placeholder(self, base_path, modified_structure_path, tmp_path):
        diff = compare(str(base_path), str(modified_structure_path))
        out = str(tmp_path / "annotated.xlsx")
        render(diff, out)
        wb = openpyxl.load_workbook(out)
        assert "(+) Audit" in wb.sheetnames
        wb.close()
