"""Tests for Flask web UI routes."""

from __future__ import annotations

import io
from pathlib import Path

import pytest
import openpyxl

from exceldiff.web.app import create_app


@pytest.fixture
def client():
    app = create_app()
    app.config["TESTING"] = True
    with app.test_client() as client:
        yield client


@pytest.fixture
def sample_xlsx_bytes():
    """Generate a minimal .xlsx file as bytes."""
    def _make(formulas=False):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Hello"
        if formulas:
            ws["B1"] = "=LEN(A1)"
        else:
            ws["B1"] = "=UPPER(A1)"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
    return _make


class TestWebRoutes:
    def test_index_returns_200(self, client):
        resp = client.get("/")
        assert resp.status_code == 200
        assert b"Compare Excel Workbooks" in resp.data

    def test_compare_requires_files(self, client):
        resp = client.post("/compare", follow_redirects=True)
        assert b"Please select" in resp.data

    def test_compare_rejects_non_xlsx(self, client):
        data = {
            "left": (io.BytesIO(b"hello"), "test.txt"),
            "right": (io.BytesIO(b"hello"), "test.txt"),
        }
        resp = client.post("/compare", data=data, content_type="multipart/form-data", follow_redirects=True)
        assert b"Only .xlsx" in resp.data

    def test_compare_succeeds(self, client, sample_xlsx_bytes):
        left = sample_xlsx_bytes(formulas=False)
        right = sample_xlsx_bytes(formulas=True)
        data = {
            "left": (left, "left.xlsx"),
            "right": (right, "right.xlsx"),
        }
        resp = client.post("/compare", data=data, content_type="multipart/form-data")
        assert resp.status_code == 302  # Redirect to results

    def test_results_page_renders(self, client, sample_xlsx_bytes):
        left = sample_xlsx_bytes(formulas=False)
        right = sample_xlsx_bytes(formulas=True)
        data = {
            "left": (left, "left.xlsx"),
            "right": (right, "right.xlsx"),
        }
        resp = client.post("/compare", data=data, content_type="multipart/form-data")
        # Follow redirect
        resp = client.get(resp.headers["Location"])
        assert resp.status_code == 200
        assert b"Comparison Results" in resp.data

    def test_download_html(self, client, sample_xlsx_bytes):
        left = sample_xlsx_bytes(formulas=False)
        right = sample_xlsx_bytes(formulas=True)
        data = {
            "left": (left, "left.xlsx"),
            "right": (right, "right.xlsx"),
        }
        resp = client.post("/compare", data=data, content_type="multipart/form-data")
        location = resp.headers["Location"]
        job_id = location.rstrip("/").split("/")[-1]
        resp = client.get(f"/results/{job_id}/download/html")
        assert resp.status_code == 200
        assert b"ExcelDiff Report" in resp.data

    def test_download_excel(self, client, sample_xlsx_bytes):
        left = sample_xlsx_bytes(formulas=False)
        right = sample_xlsx_bytes(formulas=True)
        data = {
            "left": (left, "left.xlsx"),
            "right": (right, "right.xlsx"),
        }
        resp = client.post("/compare", data=data, content_type="multipart/form-data")
        location = resp.headers["Location"]
        job_id = location.rstrip("/").split("/")[-1]
        resp = client.get(f"/results/{job_id}/download/excel")
        assert resp.status_code == 200
        # .xlsx files start with PK (zip header)
        assert resp.data[:2] == b"PK"
