"""Shared test fixtures — programmatically creates .xlsx files for testing."""

from __future__ import annotations

import pytest
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path


@pytest.fixture(scope="session")
def fixtures_dir(tmp_path_factory) -> Path:
    """Create all test fixture .xlsx files in a temporary directory."""
    d = tmp_path_factory.mktemp("fixtures")
    create_base_workbook(d / "base.xlsx")
    create_modified_formulas(d / "modified_formulas.xlsx")
    create_modified_structure(d / "modified_structure.xlsx")
    create_modified_formatting(d / "modified_formatting.xlsx")
    create_identical(d / "identical.xlsx")
    create_empty(d / "empty.xlsx")
    return d


@pytest.fixture
def base_path(fixtures_dir) -> Path:
    return fixtures_dir / "base.xlsx"


@pytest.fixture
def modified_formulas_path(fixtures_dir) -> Path:
    return fixtures_dir / "modified_formulas.xlsx"


@pytest.fixture
def modified_structure_path(fixtures_dir) -> Path:
    return fixtures_dir / "modified_structure.xlsx"


@pytest.fixture
def modified_formatting_path(fixtures_dir) -> Path:
    return fixtures_dir / "modified_formatting.xlsx"


@pytest.fixture
def identical_path(fixtures_dir) -> Path:
    return fixtures_dir / "identical.xlsx"


@pytest.fixture
def empty_path(fixtures_dir) -> Path:
    return fixtures_dir / "empty.xlsx"


def create_base_workbook(path: Path) -> None:
    """Reference workbook with known formulas and structure."""
    wb = openpyxl.Workbook()

    # Sheet 1: "Summary"
    ws1 = wb.active
    ws1.title = "Summary"
    ws1["A1"] = "Item"
    ws1["B1"] = "Value"
    ws1["A2"] = "Total"
    ws1["B2"] = "=SUM(Data!B2:B10)"
    ws1["A3"] = "Average"
    ws1["B3"] = "=AVERAGE(Data!B2:B10)"
    ws1["A4"] = "Count"
    ws1["B4"] = "=COUNTA(Data!A2:A10)"
    ws1.merge_cells("A6:B6")
    ws1["A6"] = "Report Footer"
    ws1.freeze_panes = "A2"
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 15
    ws1["A1"].comment = Comment("Header column", "Author")

    # Sheet 2: "Data"
    ws2 = wb.create_sheet("Data")
    ws2["A1"] = "Name"
    ws2["B1"] = "Amount"
    ws2["C1"] = "Markup"
    for i in range(2, 11):
        ws2[f"A{i}"] = f"Item {i - 1}"
        ws2[f"B{i}"] = i * 10
        ws2[f"C{i}"] = f"=B{i}*1.2"
    ws2["B11"] = "=SUM(B2:B10)"
    ws2["C11"] = "=SUM(C2:C10)"

    # Sheet 3: "Config"
    ws3 = wb.create_sheet("Config")
    ws3["A1"] = "VAT Rate"
    ws3["B1"] = 0.2
    ws3["A2"] = "Currency"
    ws3["B2"] = "GBP"

    wb.save(path)


def create_modified_formulas(path: Path) -> None:
    """Same structure as base, but several formulas changed."""
    wb = openpyxl.Workbook()

    # Sheet 1: "Summary" — formulas changed
    ws1 = wb.active
    ws1.title = "Summary"
    ws1["A1"] = "Item"
    ws1["B1"] = "Value"
    ws1["A2"] = "Total"
    ws1["B2"] = '=SUMIF(Data!B2:B10,">0")'  # Changed from =SUM
    ws1["A3"] = "Average"
    ws1["B3"] = "=AVERAGEIF(Data!B2:B10,\">0\")"  # Changed from =AVERAGE
    ws1["A4"] = "Count"
    ws1["B4"] = "=COUNTA(Data!A2:A10)"  # Unchanged
    ws1.merge_cells("A6:B6")
    ws1["A6"] = "Report Footer"
    ws1.freeze_panes = "A2"
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 15
    ws1["A1"].comment = Comment("Header column", "Author")

    # Sheet 2: "Data" — markup formula changed
    ws2 = wb.create_sheet("Data")
    ws2["A1"] = "Name"
    ws2["B1"] = "Amount"
    ws2["C1"] = "Markup"
    for i in range(2, 11):
        ws2[f"A{i}"] = f"Item {i - 1}"
        ws2[f"B{i}"] = i * 10
        ws2[f"C{i}"] = f"=B{i}*1.25"  # Changed from 1.2 to 1.25
    ws2["B11"] = "=SUM(B2:B10)"
    ws2["C11"] = "=SUM(C2:C10)"
    ws2["D2"] = "=B2+C2"  # New formula column

    # Sheet 3: "Config"
    ws3 = wb.create_sheet("Config")
    ws3["A1"] = "VAT Rate"
    ws3["B1"] = 0.2
    ws3["A2"] = "Currency"
    ws3["B2"] = "GBP"

    wb.save(path)


def create_modified_structure(path: Path) -> None:
    """Structural changes vs base: renamed sheet, added sheet, removed merge."""
    wb = openpyxl.Workbook()

    # Sheet 1: "Summary" — merge removed, freeze changed
    ws1 = wb.active
    ws1.title = "Summary"
    ws1["A1"] = "Item"
    ws1["B1"] = "Value"
    ws1["A2"] = "Total"
    ws1["B2"] = "=SUM(Data!B2:B10)"
    ws1["A3"] = "Average"
    ws1["B3"] = "=AVERAGE(Data!B2:B10)"
    ws1["A4"] = "Count"
    ws1["B4"] = "=COUNTA(Data!A2:A10)"
    # No merge (removed)
    ws1["A6"] = "Report Footer"
    ws1.freeze_panes = "A3"  # Changed from A2
    ws1.column_dimensions["A"].width = 25  # Changed from 20
    ws1.column_dimensions["B"].width = 15

    # Sheet 2: "Data" — same
    ws2 = wb.create_sheet("Data")
    ws2["A1"] = "Name"
    ws2["B1"] = "Amount"
    ws2["C1"] = "Markup"
    for i in range(2, 11):
        ws2[f"A{i}"] = f"Item {i - 1}"
        ws2[f"B{i}"] = i * 10
        ws2[f"C{i}"] = f"=B{i}*1.2"
    ws2["B11"] = "=SUM(B2:B10)"
    ws2["C11"] = "=SUM(C2:C10)"

    # Sheet 3: "Settings" (renamed from "Config")
    ws3 = wb.create_sheet("Settings")
    ws3["A1"] = "VAT Rate"
    ws3["B1"] = 0.2

    # Sheet 4: "Audit" (new sheet)
    ws4 = wb.create_sheet("Audit")
    ws4["A1"] = "Timestamp"
    ws4["B1"] = "Action"

    wb.save(path)


def create_modified_formatting(path: Path) -> None:
    """Same formulas and structure as base, but formatting differs."""
    wb = openpyxl.Workbook()

    # Sheet 1: "Summary" — formatting changed
    ws1 = wb.active
    ws1.title = "Summary"
    ws1["A1"] = "Item"
    ws1["A1"].font = Font(bold=True, size=14)  # Changed font
    ws1["B1"] = "Value"
    ws1["B1"].font = Font(bold=True, size=14)
    ws1["A2"] = "Total"
    ws1["B2"] = "=SUM(Data!B2:B10)"
    ws1["B2"].fill = PatternFill("solid", fgColor="FFFF00")  # Added fill
    ws1["A3"] = "Average"
    ws1["B3"] = "=AVERAGE(Data!B2:B10)"
    ws1["A4"] = "Count"
    ws1["B4"] = "=COUNTA(Data!A2:A10)"
    ws1.merge_cells("A6:B6")
    ws1["A6"] = "Report Footer"
    ws1.freeze_panes = "A2"
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 15
    ws1["A1"].comment = Comment("Header column", "Author")

    # Sheet 2: "Data"
    ws2 = wb.create_sheet("Data")
    ws2["A1"] = "Name"
    ws2["B1"] = "Amount"
    ws2["C1"] = "Markup"
    for i in range(2, 11):
        ws2[f"A{i}"] = f"Item {i - 1}"
        ws2[f"B{i}"] = i * 10
        ws2[f"C{i}"] = f"=B{i}*1.2"
    ws2["B11"] = "=SUM(B2:B10)"
    ws2["C11"] = "=SUM(C2:C10)"

    # Sheet 3: "Config"
    ws3 = wb.create_sheet("Config")
    ws3["A1"] = "VAT Rate"
    ws3["B1"] = 0.2
    ws3["A2"] = "Currency"
    ws3["B2"] = "GBP"

    wb.save(path)


def create_identical(path: Path) -> None:
    """Exact copy of the base workbook."""
    create_base_workbook(path)


def create_empty(path: Path) -> None:
    """Empty workbook with one empty sheet."""
    wb = openpyxl.Workbook()
    wb.save(path)
